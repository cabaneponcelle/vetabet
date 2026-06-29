# Extrait le planning réel depuis le .xlsm (couleur de cellule = activité)
# et produit prisma/planning-data.json consommé par import-planning.ts.
import json, re
import openpyxl


def norm_name(s):
    s = str(s).strip()
    if s.startswith("Anne So"):
        return "Anne Sophie"
    return s


def is_interim(s):
    # colonnes intérimaires : "Int 1", "In4 Elisa", etc.
    return s.startswith("Int") or bool(re.match(r"^In\d", s)) or "Int" in s

XLSM = r"C:\Users\timta\AppData\Local\Temp\planning_2025_-_assistovet_-_JUIN26 (1).xlsm"
OUT = r"C:\Users\timta\OneDrive - Haute Ecole de Namur-Liege-Luxembourg\reste\Documents\V2clinique\planning-assistovet\prisma\planning-data.json"

# Doublons sémantiques (comme build3.py) -> couleur canonique
UNIFY = {"FF0000": "FF5050", "F714E9": "FF00FF", "ED7D31": "F88628"}

# Couleur -> nom d'activité (doit matcher les noms du seed)
COLORMAP = {
    "66FFFF": "Consultation",
    "F4B183": "Consultation spécialisée",
    "FBE5D6": "Formation consult spé",
    "FF00FF": "Urgences",
    "00FF00": "Hospitalisation",
    "E2F0D9": "Formation hospi",
    "FF5050": "Chirurgie tissus mous",
    "CCCCFF": "Chirurgie spécialisée",
    "FFA6A6": "Formation chirurgie",
    "7030A0": "Anesthésie",
    "CC99FF": "Formation anesthésie",
    "3AA87E": "Échographie",
    "0070C0": "Scanner",
    "87D5B7": "Formation imagerie",
    "F88628": "Dentisterie",
    "FFCCFF": "Back up",
    "0066FF": "Bureau",
    "8FAADC": "Divers",
}


def cell_activity(cell):
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    hexv = None
    if getattr(fg, "type", None) == "rgb" and isinstance(fg.rgb, str):
        hexv = fg.rgb[-6:].upper()
    if not hexv or hexv in ("FFFFFF", "000000"):
        return None
    hexv = UNIFY.get(hexv, hexv)
    return COLORMAP.get(hexv)


def block_headers(ws):
    rows = []
    for r in range(1, ws.max_row + 2):
        v = ws.cell(r, 2).value
        if v and "Heures" in str(v):
            rows.append(r)
        if len(rows) == 7:
            break
    return rows


def main():
    wb = openpyxl.load_workbook(XLSM, data_only=True)
    out = []
    stats = {"mapped": 0, "unmapped": 0}
    unmapped_colors = {}

    for w in range(1, 5):
        ws = wb[f"semaine {w}"]
        headers = block_headers(ws)
        for d, hr in enumerate(headers):  # d = 0 (lundi) .. 6 (dimanche)
            # noms des personnes (colonnes C..)
            cols = {}
            for c in range(3, 46):
                name = ws.cell(hr, c).value
                if name is None:
                    continue
                raw = str(name).strip()
                if raw == "" or is_interim(raw):
                    continue
                cols[c] = norm_name(raw)
            for c, person in cols.items():
                # lit les 14 créneaux horaires
                slots = []
                for k in range(14):
                    cell = ws.cell(hr + 1 + k, c)
                    act = cell_activity(cell)
                    # diagnostic couleurs non mappées
                    if act is None and cell.fill and cell.fill.patternType == "solid":
                        fg = cell.fill.fgColor
                        if getattr(fg, "type", None) == "rgb" and isinstance(fg.rgb, str):
                            hx = fg.rgb[-6:].upper()
                            if hx not in ("FFFFFF", "000000"):
                                unmapped_colors[hx] = unmapped_colors.get(hx, 0) + 1
                    slots.append(act)
                # fusionne les créneaux consécutifs de même activité
                k = 0
                while k < 14:
                    if slots[k] is None:
                        k += 1
                        continue
                    act = slots[k]
                    start = k
                    while k < 14 and slots[k] == act:
                        k += 1
                    end = k  # exclusif
                    h_start = 9 + start
                    h_end = min(9 + end, 23)  # cap à 23:00 (pas de créneau de nuit)
                    if h_end <= h_start:
                        continue
                    out.append({
                        "worker": person,
                        "activity": act,
                        "week": w,
                        "day": d,
                        "heureDebut": f"{h_start:02d}:00",
                        "heureFin": f"{h_end:02d}:00",
                    })
                    stats["mapped"] += 1

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)

    print(f"Blocs créés : {len(out)}")
    print(f"Couleurs non mappées : {unmapped_colors}")
    # aperçu
    persons = sorted(set(o["worker"] for o in out))
    print(f"Personnes ({len(persons)}) : {persons}")


if __name__ == "__main__":
    main()
